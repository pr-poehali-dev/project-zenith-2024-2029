"""
Бэкенд функция: разбор Excel-файлов графика техпроцесса и выгрузки ПО Статистика.
Принимает base64-закодированный Excel-файл, возвращает распознанные данные для суточного задания или отчёта.
"""
import json
import base64
import os
import io
from datetime import date, datetime

import openpyxl
import psycopg2

SCHEMA = "t_p4383093_project_zenith_2024_"

RESPONSIBLE_KEYWORDS = ["калибровка", "ориентация", "сопротивлени", "изоляци"]
SHUTDOWN_KEYWORDS = ["калибровка", "ориентация", "сопротивлени", "изоляци"]
TWO_PERSONS_KEYWORDS = ["калибровка", "ориентация", "сопротивлени", "изоляци"]
VOICE_KEYWORDS = ["речевой информатор", "информатор"]
CALIBRATION_KEYWORDS = ["калибровка"]
ORIENTATION_KEYWORDS = ["ориентация"]
INSULATION_KEYWORDS = ["сопротивлени", "изоляци"]

CORS_HEADERS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, POST, PUT, DELETE, OPTIONS",
    "Access-Control-Allow-Headers": "Content-Type, X-User-Id, X-Auth-Token, X-Session-Id",
    "Access-Control-Max-Age": "86400",
}

def get_conn():
    return psycopg2.connect(os.environ["DATABASE_URL"])

def classify_work(work_text: str):
    t = work_text.lower()
    responsible = any(k in t for k in RESPONSIBLE_KEYWORDS)
    shutdown = any(k in t for k in SHUTDOWN_KEYWORDS)
    two_persons = any(k in t for k in TWO_PERSONS_KEYWORDS)
    voice_check = any(k in t for k in VOICE_KEYWORDS)
    calibration = any(k in t for k in CALIBRATION_KEYWORDS)
    orientation = any(k in t for k in ORIENTATION_KEYWORDS)
    insulation_check = any(k in t for k in INSULATION_KEYWORDS)
    return responsible, shutdown, two_persons, voice_check, calibration, orientation, insulation_check

def _find_columns(rows):
    """Ищет индексы колонок в заголовочной строке."""
    col_device = col_section = col_work = col_planned = col_date = None
    header_row = 0
    for i, row in enumerate(rows[:15]):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        for j, cell in enumerate(row_lower):
            if "устройств" in cell:
                col_device = j
            if "участок" in cell or "секция" in cell:
                col_section = j
            if ("работ" in cell and "перечень" in cell) or "наименование работ" in cell or "вид работ" in cell:
                col_work = j
            if "продолжительност" in cell or ("план" in cell and "дат" not in cell):
                col_planned = j
            if "дата" in cell:
                col_date = j
        if col_device is not None:
            header_row = i
            break
    if col_device is None:
        col_device, col_section, col_work, col_planned = 0, 1, 2, 3
    return header_row, col_device, col_section, col_work, col_planned, col_date

def parse_schedule_excel(wb: openpyxl.Workbook, target_date: date):
    """Парсит график техпроцесса, ищет работы на указанную дату."""
    tasks = []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return tasks

    header_row, col_device, col_section, col_work, col_planned, col_date = _find_columns(rows)

    for row in rows[header_row + 1:]:
        if not any(row):
            continue
        device = str(row[col_device]).strip() if col_device is not None and row[col_device] else ""
        section = str(row[col_section]).strip() if col_section is not None and row[col_section] else ""
        work = str(row[col_work]).strip() if col_work is not None and row[col_work] else ""
        planned = str(row[col_planned]).strip() if col_planned is not None and row[col_planned] else ""

        if not device or not work or device == "None":
            continue

        row_date = None
        if col_date is not None and row[col_date]:
            val = row[col_date]
            if isinstance(val, (datetime, date)):
                row_date = val.date() if isinstance(val, datetime) else val
            else:
                for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        row_date = datetime.strptime(str(val).strip(), fmt).date()
                        break
                    except Exception:
                        pass

        if col_date is not None and row_date and row_date != target_date:
            continue

        resp, shut, two, voice, cal, ori, ins = classify_work(work)
        tasks.append({
            "device": device,
            "section": section,
            "work": work,
            "planned_duration": planned or "—",
            "responsible": resp,
            "shutdown": shut,
            "two_persons": two,
            "voice_check": voice,
            "calibration": cal,
            "orientation": ori,
            "insulation_check": ins,
            "executor": "",
            "order_number": "",
        })
    return tasks

def parse_schedule_excel_bulk(wb: openpyxl.Workbook, year: int, month: int):
    """
    Парсит весь оперативный план за указанный год и месяц.
    Возвращает dict: {date: [tasks]} сгруппированный по датам.
    Если колонки дат нет — все записи относятся к первому дню периода.
    """
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}

    header_row, col_device, col_section, col_work, col_planned, col_date = _find_columns(rows)

    from calendar import monthrange
    days_in_month = monthrange(year, month)[1]
    fallback_date = date(year, month, 1)

    tasks_by_date: dict = {}

    for row in rows[header_row + 1:]:
        if not any(row):
            continue
        device = str(row[col_device]).strip() if col_device is not None and row[col_device] else ""
        section = str(row[col_section]).strip() if col_section is not None and row[col_section] else ""
        work = str(row[col_work]).strip() if col_work is not None and row[col_work] else ""
        planned = str(row[col_planned]).strip() if col_planned is not None and row[col_planned] else ""

        if not device or not work or device == "None":
            continue

        row_date = None
        if col_date is not None and row[col_date]:
            val = row[col_date]
            if isinstance(val, (datetime, date)):
                row_date = val.date() if isinstance(val, datetime) else val
            else:
                for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                    try:
                        row_date = datetime.strptime(str(val).strip(), fmt).date()
                        break
                    except Exception:
                        pass

        # Фильтруем: только нужный год и месяц
        if row_date:
            if row_date.year != year or row_date.month != month:
                continue
        else:
            row_date = fallback_date

        resp, shut, two, voice, cal, ori, ins = classify_work(work)
        task = {
            "device": device,
            "section": section,
            "work": work,
            "planned_duration": planned or "—",
            "responsible": resp,
            "shutdown": shut,
            "two_persons": two,
            "voice_check": voice,
            "calibration": cal,
            "orientation": ori,
            "insulation_check": ins,
            "executor": "",
            "order_number": "",
        }
        tasks_by_date.setdefault(str(row_date), []).append(task)

    return tasks_by_date

def parse_statistics_excel(wb: openpyxl.Workbook):
    """Парсит выгрузку ПО Статистика. Возвращает список устройств с фактическими данными."""
    records = []
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return records

    header_row = 0
    col_device = col_duration = col_door = col_calibration = col_cal_result = col_shutdown = None
    for i, row in enumerate(rows[:10]):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        for j, cell in enumerate(row_lower):
            if "устройств" in cell:
                col_device = j
            if "продолжительност" in cell or "время" in cell or "длительност" in cell:
                col_duration = j
            if "дверь" in cell or "открыт" in cell:
                col_door = j
            if "калибровк" in cell:
                col_calibration = j
            if "результат" in cell and col_calibration is not None:
                col_cal_result = j
            if "отказ" in cell or "офлайн" in cell or "offline" in cell:
                col_shutdown = j
        if col_device is not None:
            header_row = i
            break

    if col_device is None:
        col_device = 0

    for row in rows[header_row + 1:]:
        if not any(row):
            continue
        device = str(row[col_device]).strip() if row[col_device] else ""
        if not device or device == "None":
            continue

        duration = str(row[col_duration]).strip() if col_duration and row[col_duration] else "—"
        door_val = str(row[col_door]).strip().lower() if col_door and row[col_door] else ""
        staff_present = "откр" in door_val or "1" in door_val or "да" in door_val or "true" in door_val
        cal_val = str(row[col_calibration]).strip().lower() if col_calibration and row[col_calibration] else ""
        calibration_done = "выполн" in cal_val or "1" in cal_val or "да" in cal_val or "true" in cal_val
        cal_result = str(row[col_cal_result]).strip() if col_cal_result and row[col_cal_result] else ("Выполнена" if calibration_done else "Не выполнена")
        shut_val = str(row[col_shutdown]).strip().lower() if col_shutdown and row[col_shutdown] else ""
        shutdown_fact = "отказ" in shut_val or "офлайн" in shut_val or "offline" in shut_val or "1" in shut_val

        records.append({
            "device": device,
            "staff_present": staff_present,
            "actual_duration": duration,
            "calibration_done": calibration_done,
            "calibration_result": cal_result,
            "shutdown_fact": shutdown_fact,
        })
    return records

def save_tasks(tasks: list, task_date: date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {SCHEMA}.esp_daily_tasks WHERE task_date = %s", (task_date,))
    for t in tasks:
        cur.execute(f"""
            INSERT INTO {SCHEMA}.esp_daily_tasks
            (task_date, device, section, work, planned_duration, responsible, shutdown, two_persons,
             voice_check, calibration, orientation, insulation_check)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            task_date, t["device"], t["section"], t["work"], t["planned_duration"],
            t["responsible"], t["shutdown"], t["two_persons"], t["voice_check"],
            t["calibration"], t["orientation"], t["insulation_check"]
        ))
    conn.commit()
    cur.close()
    conn.close()

def save_report(records: list, task_date: date, tasks: list):
    """Сохраняет отчёт, сравнивая с плановыми данными задания."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {SCHEMA}.esp_reports WHERE report_date = %s", (task_date,))
    plan_map = {t["device"]: t for t in tasks}
    for r in records:
        plan = plan_map.get(r["device"], {})
        planned_str = plan.get("planned_duration", "")
        matches_plan = True
        deviation_notes = []
        if planned_str and planned_str != "—" and r["actual_duration"] and r["actual_duration"] != "—":
            try:
                def to_minutes(s):
                    parts = s.strip().split(":")
                    return int(parts[0]) * 60 + int(parts[1])
                planned_m = to_minutes(planned_str)
                actual_m = to_minutes(r["actual_duration"])
                diff = abs(actual_m - planned_m)
                if diff > 15:
                    matches_plan = False
                    deviation_notes.append(f"Отклонение по времени: план {planned_str}, факт {r['actual_duration']}")
            except Exception:
                pass
        if not r["staff_present"]:
            matches_plan = False
            deviation_notes.append("Персонал отсутствовал")
        cur.execute(f"""
            INSERT INTO {SCHEMA}.esp_reports
            (report_date, device, staff_present, actual_duration, matches_plan,
             calibration_done, calibration_result, shutdown_fact, deviation_notes)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            task_date, r["device"], r["staff_present"], r["actual_duration"],
            matches_plan, r["calibration_done"], r["calibration_result"],
            r["shutdown_fact"], "; ".join(deviation_notes)
        ))
    conn.commit()
    cur.close()
    conn.close()

def generate_sample_plan(year: int, month: int) -> bytes:
    """Генерирует пример оперативного плана на месяц в формате Excel."""
    from calendar import monthrange
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Оперативный план"

    headers = ["Дата", "Устройство", "Участок", "Перечень работ", "Плановая продолжительность"]
    ws.append(headers)

    bold = openpyxl.styles.Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    devices = [
        ("АЛС-1 (ПК 12+450)", "Участок №1"),
        ("АЛС-2 (ПК 18+200)", "Участок №1"),
        ("САУТ-Ц (ст. Северная)", "Участок №2"),
        ("ТСКБМ (ПК 24+100)", "Участок №2"),
        ("КЛУБ-У (ст. Южная)", "Участок №3"),
        ("УКСПС (ПК 30+780)", "Участок №3"),
    ]
    # Разнообразные работы — для наглядного срабатывания всех признаков:
    # калибровка / ориентация / изоляция — ответственные, требуют выключения, в два лица
    # речевой информатор — отдельный признак
    works = [
        ("Калибровка ПУ тракта", "01:30"),                          # ответственная, выключение, два лица, калибровка
        ("Ориентация антенны напольного устройства", "02:00"),      # ответственная, выключение, два лица, ориентация
        ("Проверка сопротивления изоляции жил кабеля", "00:45"),    # ответственная, выключение, два лица, изоляция
        ("Проверка работы речевого информатора", "00:30"),          # речевой информатор
        ("Техническое обслуживание ТО-2", "01:15"),                 # обычная
        ("Внешний осмотр и чистка аппаратуры", "00:40"),            # обычная
        ("Проверка показаний и анализ работы устройства", "00:50"), # обычная
    ]

    days = monthrange(year, month)[1]
    for day in range(1, days + 1):
        d = date(year, month, day)
        # Каждый рабочий день — несколько устройств с разными работами,
        # смещение обеспечивает чередование всех типов работ по дням
        for idx, (device, section) in enumerate(devices):
            work, planned = works[(day + idx) % len(works)]
            ws.append([d.strftime("%d.%m.%Y"), device, section, work, planned])

    for col, width in zip("ABCDE", [14, 26, 16, 40, 28]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

TASK_COLS = ["id","device","section","work","planned_duration","responsible","shutdown",
             "two_persons","voice_check","calibration","orientation","insulation_check",
             "executor","order_number","tech_card","location","done","transfer_date",
             "transfer_reason","car_owner","fuel_spent","transport_type","arrival_time",
             "departure_time","power_off_time","power_on_time","total_off_hours"]

def get_tasks(task_date: date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, device, section, work, planned_duration, responsible, shutdown, two_persons,
               voice_check, calibration, orientation, insulation_check, executor, order_number,
               tech_card, location, done, transfer_date, transfer_reason, car_owner, fuel_spent,
               transport_type, arrival_time, departure_time, power_off_time, power_on_time, total_off_hours
        FROM {SCHEMA}.esp_daily_tasks WHERE task_date = %s ORDER BY id
    """, (task_date,))
    rows = []
    for row in cur.fetchall():
        d = dict(zip(TASK_COLS, row))
        if d.get("transfer_date"):
            d["transfer_date"] = str(d["transfer_date"])
        rows.append(d)
    cur.close()
    conn.close()
    return rows

UNPLANNED_COLS = ["id","device","location","executor","power_off_time","power_on_time",
                  "total_off_hours","is_failure","is_pre_failure","reason"]

def get_unplanned_trips(trip_date: date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, device, location, executor, power_off_time, power_on_time,
               total_off_hours, is_failure, is_pre_failure, reason
        FROM {SCHEMA}.esp_unplanned_trips WHERE trip_date = %s ORDER BY id
    """, (trip_date,))
    rows = [dict(zip(UNPLANNED_COLS, row)) for row in cur.fetchall()]
    cur.close()
    conn.close()
    return rows

def add_unplanned_trip(trip_date: date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        INSERT INTO {SCHEMA}.esp_unplanned_trips (trip_date) VALUES (%s) RETURNING id
    """, (trip_date,))
    new_id = cur.fetchone()[0]
    conn.commit()
    cur.close()
    conn.close()
    return new_id

def update_unplanned_trip(trip_id: int, field: str, value: str):
    allowed = {"device","location","executor","power_off_time","power_on_time",
               "total_off_hours","is_failure","is_pre_failure","reason"}
    if field not in allowed:
        return
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.esp_unplanned_trips SET {field}=%s WHERE id=%s", (value, trip_id))
    conn.commit()
    cur.close()
    conn.close()

def delete_unplanned_trip(trip_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"DELETE FROM {SCHEMA}.esp_unplanned_trips WHERE id=%s", (trip_id,))
    conn.commit()
    cur.close()
    conn.close()

def update_task_field(task_id: int, field: str, value):
    allowed = {"executor","order_number","tech_card","location","done","transfer_reason",
               "car_owner","fuel_spent","transport_type","arrival_time","departure_time",
               "power_off_time","power_on_time","total_off_hours"}
    if field not in allowed:
        return
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"UPDATE {SCHEMA}.esp_daily_tasks SET {field}=%s WHERE id=%s", (value, task_id))
    conn.commit()
    cur.close()
    conn.close()

def transfer_task(task_id: int, new_date: date, reason: str):
    """Переносит работу на другой день."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        UPDATE {SCHEMA}.esp_daily_tasks
        SET task_date=%s, transfer_date=%s, transfer_reason=%s
        WHERE id=%s
    """, (new_date, new_date, reason, task_id))
    conn.commit()
    cur.close()
    conn.close()

def get_report(task_date: date):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT id, device, staff_present, actual_duration, matches_plan,
               calibration_done, calibration_result, shutdown_fact, deviation_notes
        FROM {SCHEMA}.esp_reports WHERE report_date = %s ORDER BY id
    """, (task_date,))
    cols = ["id","device","staff_present","actual_duration","matches_plan",
            "calibration_done","calibration_result","shutdown_fact","deviation_notes"]
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]
    cur.close()
    conn.close()
    return rows

def get_monthly_report(year: int, month: int):
    """Сводный месячный отчёт: агрегация по устройствам за все сутки месяца."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        SELECT
            r.device,
            COUNT(DISTINCT r.report_date) AS days_worked,
            COUNT(r.id) AS total_records,
            SUM(CASE WHEN r.matches_plan THEN 1 ELSE 0 END) AS matches_plan_count,
            SUM(CASE WHEN NOT r.matches_plan THEN 1 ELSE 0 END) AS deviations_count,
            SUM(CASE WHEN r.staff_present THEN 1 ELSE 0 END) AS staff_present_days,
            SUM(CASE WHEN r.calibration_done THEN 1 ELSE 0 END) AS calibrations_done,
            SUM(CASE WHEN r.shutdown_fact THEN 1 ELSE 0 END) AS shutdowns_count,
            array_agg(DISTINCT r.report_date ORDER BY r.report_date) AS work_dates,
            array_agg(r.deviation_notes) FILTER (WHERE r.deviation_notes != '') AS all_deviations
        FROM {SCHEMA}.esp_reports r
        WHERE EXTRACT(YEAR FROM r.report_date) = %s
          AND EXTRACT(MONTH FROM r.report_date) = %s
        GROUP BY r.device
        ORDER BY r.device
    """, (year, month))
    cols = ["device","days_worked","total_records","matches_plan_count","deviations_count",
            "staff_present_days","calibrations_done","shutdowns_count","work_dates","all_deviations"]
    rows = []
    for row in cur.fetchall():
        d = dict(zip(cols, row))
        d["work_dates"] = [str(x) for x in (d["work_dates"] or [])]
        d["all_deviations"] = [x for x in (d["all_deviations"] or []) if x]
        total = d["total_records"] or 1
        d["plan_percent"] = round(d["matches_plan_count"] / total * 100)
        rows.append(d)

    cur.execute(f"""
        SELECT COUNT(DISTINCT report_date) FROM {SCHEMA}.esp_reports
        WHERE EXTRACT(YEAR FROM report_date) = %s AND EXTRACT(MONTH FROM report_date) = %s
    """, (year, month))
    total_days = cur.fetchone()[0] or 0

    cur.execute(f"""
        SELECT COUNT(*) FROM {SCHEMA}.esp_reports
        WHERE EXTRACT(YEAR FROM report_date) = %s AND EXTRACT(MONTH FROM report_date) = %s
          AND NOT matches_plan
    """, (year, month))
    total_deviations = cur.fetchone()[0] or 0

    cur.close()
    conn.close()
    return {"rows": rows, "total_days": total_days, "total_deviations": total_deviations}

def update_task_executor(task_id: int, executor: str, order_number: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(f"""
        UPDATE {SCHEMA}.esp_daily_tasks SET executor=%s, order_number=%s WHERE id=%s
    """, (executor, order_number, task_id))
    conn.commit()
    cur.close()
    conn.close()

def handler(event: dict, context) -> dict:
    if event.get("httpMethod") == "OPTIONS":
        return {"statusCode": 200, "headers": CORS_HEADERS, "body": ""}

    method = event.get("httpMethod", "GET")
    path = event.get("path", "/")
    body = {}
    if event.get("body"):
        try:
            body = json.loads(event["body"])
        except Exception:
            pass

    qs = event.get("queryStringParameters") or {}
    action = qs.get("action", "")

    # GET ?action=tasks&date=2026-06-19
    if method == "GET" and action == "tasks":
        raw_date = qs.get("date", str(date.today()))
        try:
            task_date = date.fromisoformat(raw_date)
        except Exception:
            task_date = date.today()
        tasks = get_tasks(task_date)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"tasks": tasks, "date": str(task_date)})}

    # GET ?action=report&date=2026-06-19
    if method == "GET" and action == "report":
        raw_date = qs.get("date", str(date.today()))
        try:
            task_date = date.fromisoformat(raw_date)
        except Exception:
            task_date = date.today()
        report = get_report(task_date)
        tasks = get_tasks(task_date)
        deviations = sum(1 for r in report if not r["matches_plan"])
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"report": report, "tasks": tasks, "deviations": deviations, "date": str(task_date)})}

    # GET ?action=sample-plan&year=2026&month=6 — скачать пример оперативного плана
    if method == "GET" and action == "sample-plan":
        year = int(qs.get("year", str(date.today().year)))
        month = int(qs.get("month", str(date.today().month)))
        file_bytes = generate_sample_plan(year, month)
        file_b64 = base64.b64encode(file_bytes).decode("utf-8")
        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Content-Disposition": f'attachment; filename="operativnyy_plan_{year}_{month:02d}.xlsx"',
                "Access-Control-Allow-Origin": "*",
            },
            "isBase64Encoded": True,
            "body": file_b64,
        }

    # POST ?action=parse-schedule-bulk — загрузка оперативного плана за год/месяц целиком
    if method == "POST" and action == "parse-schedule-bulk":
        file_b64 = body.get("file")
        year = int(body.get("year", date.today().year))
        month = int(body.get("month", date.today().month))
        if not file_b64:
            return {"statusCode": 400, "headers": CORS_HEADERS, "body": json.dumps({"error": "Файл не передан"})}
        try:
            file_bytes = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            tasks_by_date = parse_schedule_excel_bulk(wb, year, month)
            total = 0
            for d_str, tasks in tasks_by_date.items():
                task_date = date.fromisoformat(d_str)
                save_tasks(tasks, task_date)
                total += len(tasks)
            return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"},
                    "body": json.dumps({"days": len(tasks_by_date), "total_tasks": total, "year": year, "month": month})}
        except Exception as e:
            return {"statusCode": 500, "headers": CORS_HEADERS, "body": json.dumps({"error": str(e)})}

    # POST ?action=parse-schedule — разбор графика техпроцесса
    if method == "POST" and action == "parse-schedule":
        file_b64 = body.get("file")
        raw_date = body.get("date", str(date.today()))
        if not file_b64:
            return {"statusCode": 400, "headers": CORS_HEADERS, "body": json.dumps({"error": "Файл не передан"})}
        try:
            task_date = date.fromisoformat(raw_date)
        except Exception:
            task_date = date.today()
        try:
            file_bytes = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            tasks = parse_schedule_excel(wb, task_date)
            if tasks:
                save_tasks(tasks, task_date)
            return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"tasks": tasks, "count": len(tasks), "date": str(task_date)})}
        except Exception as e:
            return {"statusCode": 500, "headers": CORS_HEADERS, "body": json.dumps({"error": str(e)})}

    # POST ?action=parse-statistics — разбор выгрузки ПО Статистика
    if method == "POST" and action == "parse-statistics":
        file_b64 = body.get("file")
        raw_date = body.get("date", str(date.today()))
        if not file_b64:
            return {"statusCode": 400, "headers": CORS_HEADERS, "body": json.dumps({"error": "Файл не передан"})}
        try:
            task_date = date.fromisoformat(raw_date)
        except Exception:
            task_date = date.today()
        try:
            file_bytes = base64.b64decode(file_b64)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            records = parse_statistics_excel(wb)
            tasks = get_tasks(task_date)
            if records:
                save_report(records, task_date, tasks)
            return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"records": records, "count": len(records), "date": str(task_date)})}
        except Exception as e:
            return {"statusCode": 500, "headers": CORS_HEADERS, "body": json.dumps({"error": str(e)})}

    # GET ?action=monthly-report&year=2026&month=6
    if method == "GET" and action == "monthly-report":
        year = int(qs.get("year", str(date.today().year)))
        month = int(qs.get("month", str(date.today().month)))
        data = get_monthly_report(year, month)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps(data)}

    # PUT ?action=update-task&id=5 — обновить любое поле задания
    if method == "PUT" and action == "update-task":
        task_id = int(qs.get("id", "0"))
        # Старый формат — ФИО и приказ сразу
        if "executor" in body and "order_number" in body and len(body) == 2:
            update_task_executor(task_id, body.get("executor", ""), body.get("order_number", ""))
        else:
            for field, value in body.items():
                update_task_field(task_id, field, value)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"ok": True})}

    # PUT ?action=transfer-task&id=5 — перенести работу на другой день
    if method == "PUT" and action == "transfer-task":
        task_id = int(qs.get("id", "0"))
        new_date_raw = body.get("new_date", "")
        reason = body.get("reason", "")
        try:
            new_date = date.fromisoformat(new_date_raw)
        except Exception:
            return {"statusCode": 400, "headers": CORS_HEADERS, "body": json.dumps({"error": "Некорректная дата переноса"})}
        transfer_task(task_id, new_date, reason)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"ok": True, "new_date": str(new_date)})}

    # GET ?action=unplanned&date=2026-06-19 — список внеплановых выездов
    if method == "GET" and action == "unplanned":
        raw_date = qs.get("date", str(date.today()))
        try:
            trip_date = date.fromisoformat(raw_date)
        except Exception:
            trip_date = date.today()
        trips = get_unplanned_trips(trip_date)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"trips": trips, "date": str(trip_date)})}

    # POST ?action=add-unplanned&date=2026-06-19 — добавить пустую строку выезда
    if method == "POST" and action == "add-unplanned":
        raw_date = qs.get("date", str(date.today()))
        try:
            trip_date = date.fromisoformat(raw_date)
        except Exception:
            trip_date = date.today()
        new_id = add_unplanned_trip(trip_date)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"id": new_id})}

    # PUT ?action=update-unplanned&id=5 — обновить поле внепланового выезда
    if method == "PUT" and action == "update-unplanned":
        trip_id = int(qs.get("id", "0"))
        for field, value in body.items():
            update_unplanned_trip(trip_id, field, value)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"ok": True})}

    # DELETE ?action=delete-unplanned&id=5 — удалить выезд
    if method == "DELETE" and action == "delete-unplanned":
        trip_id = int(qs.get("id", "0"))
        delete_unplanned_trip(trip_id)
        return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"ok": True})}

    return {"statusCode": 200, "headers": {**CORS_HEADERS, "Content-Type": "application/json"}, "body": json.dumps({"status": "Диспетчер ЭСП API работает"})}